import openpyxl
from random import uniform

"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Sheet1']

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value= numero_pedido
    planilha1.cell(linha, 2).value= 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value=preco

pedidos.save('nova_planilha.xlsx')
"""


# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

# for linha in planilha1:
#     for coluna in linha:
#         print(coluna.value)




planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha 1', 0)
planilha.create_sheet('Planilha 2', 0)

planilha1 = planilha['Planilha 1']
planilha2 = planilha['Planilha 2']


for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value= numero_pedido
    planilha1.cell(linha, 2).value= 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value=preco


for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Alvaro {linha} {round(uniform(10,100), 2)}'
    planilha2.cell(linha, 2).value = f'Sousa {linha} {round(uniform(10,100), 2)}'
    planilha2.cell(linha, 3).value = f'Oliveira {linha} {round(uniform(10,100), 2)}'


planilha.save('outra_planilha.xlsx')